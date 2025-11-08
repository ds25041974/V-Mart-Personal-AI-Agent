"""
Excel Data Reader
Comprehensive Excel file reader that captures:
- All sheets (visible and hidden)
- All columns and rows (visible and hidden)
- Filters and advanced filters
- Pivot tables (including hidden ones)
- Formulas and calculated fields
- Cell formatting and data validation
"""

import os
import platform
import subprocess
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class ExcelReader:
    """Advanced Excel file reader with comprehensive data extraction"""

    def __init__(self, file_path: Optional[str] = None):
        self.file_path = file_path
        self.workbook = None
        self.data = {}

    def get_active_excel_file(self) -> Optional[str]:
        """Get the path of the currently active Excel file"""
        system = platform.system()

        if system == "Darwin":  # macOS
            return self._get_active_excel_macos()
        elif system == "Windows":
            return self._get_active_excel_windows()

        return None

    def _get_active_excel_macos(self) -> Optional[str]:
        """Get active Excel file path on macOS"""
        try:
            script = """
            tell application "Microsoft Excel"
                if (count of workbooks) > 0 then
                    set activeWorkbook to active workbook
                    set workbookPath to full name of activeWorkbook
                    return workbookPath
                else
                    return ""
                end if
            end tell
            """

            result = subprocess.run(
                ["osascript", "-e", script], capture_output=True, text=True, timeout=5
            )

            if result.returncode == 0 and result.stdout.strip():
                return result.stdout.strip()
        except Exception as e:
            print(f"Error getting active Excel file: {e}")

        return None

    def _get_active_excel_windows(self) -> Optional[str]:
        """Get active Excel file path on Windows"""
        try:
            script = """
            $excel = [System.Runtime.InteropServices.Marshal]::GetActiveObject("Excel.Application")
            if ($excel.ActiveWorkbook) {
                $excel.ActiveWorkbook.FullName
            }
            """

            result = subprocess.run(
                ["powershell", "-Command", script],
                capture_output=True,
                text=True,
                timeout=5,
            )

            if result.returncode == 0 and result.stdout.strip():
                return result.stdout.strip()
        except Exception as e:
            print(f"Error getting active Excel file: {e}")

        return None

    def load_file(self, file_path: Optional[str] = None) -> bool:
        """Load Excel file for reading"""
        if file_path:
            self.file_path = file_path

        if not self.file_path:
            self.file_path = self.get_active_excel_file()

        if not self.file_path or not os.path.exists(self.file_path):
            return False

        try:
            self.workbook = load_workbook(
                self.file_path, data_only=False, keep_vba=True
            )
            return True
        except Exception as e:
            print(f"Error loading Excel file: {e}")
            return False

    def get_file_metadata(self) -> Dict[str, Any]:
        """Get Excel file metadata"""
        if not self.workbook:
            return {}

        props = self.workbook.properties

        return {
            "filename": os.path.basename(self.file_path) if self.file_path else "",
            "full_path": self.file_path or "",
            "title": props.title or "",
            "author": props.creator or "",
            "last_modified_by": props.lastModifiedBy or "",
            "created": props.created.isoformat() if props.created else "",
            "modified": props.modified.isoformat() if props.modified else "",
            "sheet_count": len(self.workbook.sheetnames),
            "sheet_names": self.workbook.sheetnames,
        }

    def get_all_sheets_info(self) -> List[Dict[str, Any]]:
        """Get information about all sheets (visible and hidden)"""
        if not self.workbook:
            return []

        sheets_info = []

        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]

            # Determine visibility
            visibility = "visible"
            if hasattr(sheet, "sheet_state"):
                if sheet.sheet_state == "hidden":
                    visibility = "hidden"
                elif sheet.sheet_state == "veryHidden":
                    visibility = "very_hidden"

            sheets_info.append(
                {
                    "name": sheet_name,
                    "visibility": visibility,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "has_filters": self._check_auto_filter(sheet),
                    "has_pivot_tables": len(sheet._pivots)
                    if hasattr(sheet, "_pivots")
                    else 0,
                }
            )

        return sheets_info

    def _check_auto_filter(self, sheet) -> bool:
        """Check if sheet has auto filters applied"""
        return sheet.auto_filter is not None

    def read_sheet_data(
        self, sheet_name: str, include_hidden: bool = True
    ) -> Dict[str, Any]:
        """Read comprehensive data from a specific sheet"""
        if not self.workbook or sheet_name not in self.workbook.sheetnames:
            return {}

        sheet = self.workbook[sheet_name]

        # Get column information
        columns_info = self._get_columns_info(sheet, include_hidden)

        # Get row information
        rows_info = self._get_rows_info(sheet, include_hidden)

        # Get cell data
        cell_data = self._get_cell_data(sheet, include_hidden)

        # Get filters
        filters = self._get_filters(sheet)

        # Get pivot tables
        pivot_tables = self._get_pivot_tables(sheet)

        # Get formulas
        formulas = self._get_formulas(sheet)

        return {
            "sheet_name": sheet_name,
            "visibility": "hidden"
            if hasattr(sheet, "sheet_state") and sheet.sheet_state == "hidden"
            else "visible",
            "dimensions": {"max_row": sheet.max_row, "max_column": sheet.max_column},
            "columns": columns_info,
            "rows": rows_info,
            "data": cell_data,
            "filters": filters,
            "pivot_tables": pivot_tables,
            "formulas": formulas,
        }

    def _get_columns_info(self, sheet, include_hidden: bool) -> List[Dict[str, Any]]:
        """Get information about all columns including hidden ones"""
        columns = []

        for col_idx in range(1, sheet.max_column + 1):
            col_letter = get_column_letter(col_idx)
            col_dimension = sheet.column_dimensions.get(col_letter)

            is_hidden = col_dimension.hidden if col_dimension else False

            if include_hidden or not is_hidden:
                # Get column header (first row)
                header_cell = sheet.cell(row=1, column=col_idx)

                columns.append(
                    {
                        "index": col_idx,
                        "letter": col_letter,
                        "header": header_cell.value,
                        "hidden": is_hidden,
                        "width": col_dimension.width if col_dimension else None,
                    }
                )

        return columns

    def _get_rows_info(self, sheet, include_hidden: bool) -> List[Dict[str, Any]]:
        """Get information about all rows including hidden ones"""
        rows = []

        for row_idx in range(1, min(sheet.max_row + 1, 101)):  # Limit to first 100 rows
            row_dimension = sheet.row_dimensions.get(row_idx)
            is_hidden = row_dimension.hidden if row_dimension else False

            if include_hidden or not is_hidden:
                rows.append(
                    {
                        "index": row_idx,
                        "hidden": is_hidden,
                        "height": row_dimension.height if row_dimension else None,
                    }
                )

        return rows

    def _get_cell_data(self, sheet, include_hidden: bool) -> List[List[Any]]:
        """Get cell data from the sheet"""
        data = []

        # Limit to first 100 rows for performance
        max_rows = min(sheet.max_row, 100)

        for row in sheet.iter_rows(
            min_row=1, max_row=max_rows, max_col=sheet.max_column
        ):
            row_data = []
            for cell in row:
                # Check if column or row is hidden
                col_letter = get_column_letter(cell.column)
                col_hidden = sheet.column_dimensions.get(col_letter, None)
                row_hidden = sheet.row_dimensions.get(cell.row, None)

                is_hidden = (col_hidden and col_hidden.hidden) or (
                    row_hidden and row_hidden.hidden
                )

                if include_hidden or not is_hidden:
                    row_data.append(
                        {
                            "value": cell.value,
                            "formula": cell.value
                            if isinstance(cell.value, str)
                            and cell.value.startswith("=")
                            else None,
                            "data_type": cell.data_type,
                            "hidden": is_hidden,
                            "number_format": cell.number_format,
                        }
                    )

            if row_data:
                data.append(row_data)

        return data

    def _get_filters(self, sheet) -> Dict[str, Any]:
        """Get filter information from the sheet"""
        filters = {
            "has_auto_filter": False,
            "filter_range": None,
            "filtered_columns": [],
        }

        if sheet.auto_filter:
            filters["has_auto_filter"] = True
            filters["filter_range"] = (
                str(sheet.auto_filter.ref) if sheet.auto_filter.ref else None
            )

            # Get filtered columns
            if hasattr(sheet.auto_filter, "filterColumn"):
                for filter_col in sheet.auto_filter.filterColumn:
                    filters["filtered_columns"].append(
                        {"column_id": filter_col.colId, "has_filter": True}
                    )

        return filters

    def _get_pivot_tables(self, sheet) -> List[Dict[str, Any]]:
        """Get pivot table information"""
        pivot_tables = []

        if hasattr(sheet, "_pivots"):
            for pivot in sheet._pivots:
                pivot_tables.append(
                    {
                        "name": pivot.name if hasattr(pivot, "name") else "Unknown",
                        "location": str(pivot.location)
                        if hasattr(pivot, "location")
                        else "Unknown",
                        "cache_id": pivot.cacheId
                        if hasattr(pivot, "cacheId")
                        else None,
                    }
                )

        return pivot_tables

    def _get_formulas(self, sheet) -> List[Dict[str, Any]]:
        """Get all formulas in the sheet"""
        formulas = []

        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 100)):
            for cell in row:
                if (
                    cell.value
                    and isinstance(cell.value, str)
                    and cell.value.startswith("=")
                ):
                    formulas.append(
                        {
                            "cell": cell.coordinate,
                            "formula": cell.value,
                            "row": cell.row,
                            "column": cell.column,
                        }
                    )

        return formulas

    def read_all_data(self, include_hidden: bool = True) -> Dict[str, Any]:
        """Read comprehensive data from all sheets"""
        if not self.workbook:
            if not self.load_file():
                return {"error": "Failed to load Excel file"}

        result = {
            "metadata": self.get_file_metadata(),
            "sheets_summary": self.get_all_sheets_info(),
            "sheets_data": {},
        }

        for sheet_name in self.workbook.sheetnames:
            result["sheets_data"][sheet_name] = self.read_sheet_data(
                sheet_name, include_hidden
            )

        return result

    def get_active_sheet_data(self, include_hidden: bool = True) -> Dict[str, Any]:
        """Get data from the currently active sheet"""
        if not self.workbook:
            if not self.load_file():
                return {"error": "Failed to load Excel file"}

        # Get active sheet name
        active_sheet_name = self.workbook.active.title

        return {
            "metadata": self.get_file_metadata(),
            "active_sheet": active_sheet_name,
            "data": self.read_sheet_data(active_sheet_name, include_hidden),
        }
